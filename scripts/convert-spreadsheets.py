#!/usr/bin/env python3
"""
Convert XLSX spreadsheets to JSON format for the survival app.
Reads all xlsx files from spreadsheets/ directory and converts them to JSON.
Each spreadsheet becomes one JSON file with all its sheets.
"""

import os
import json
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

def convert_xlsx_to_json(xlsx_path):
    """
    Convert a single XLSX file to JSON format with all sheets.

    Args:
        xlsx_path: Path to the XLSX file

    Returns:
        dict with filename and sheets data
    """
    workbook = load_workbook(xlsx_path)

    sheets_data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Extract data from sheet
        rows = []
        headers = None

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                # First row is headers
                headers = [str(cell).strip() if cell else "" for cell in row]
            else:
                # Convert row to dict using headers
                row_dict = {}
                for col_idx, cell_value in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        # Convert None to null, handle datetime
                        if cell_value is None:
                            row_dict[header] = None
                        elif isinstance(cell_value, datetime):
                            row_dict[header] = cell_value.isoformat()
                        else:
                            row_dict[header] = cell_value

                rows.append(row_dict)

        sheets_data[sheet_name] = {
            "headers": headers,
            "rows": rows,
            "rowCount": len(rows)
        }

    workbook.close()

    return sheets_data

def main():
    """Main conversion process."""

    # Define paths
    app_root = Path(__file__).parent.parent
    spreadsheets_dir = app_root / "spreadsheets"
    output_dir = app_root / "data" / "spreadsheets"

    # Create output directory
    output_dir.mkdir(parents=True, exist_ok=True)

    # Find all xlsx files
    xlsx_files = sorted(spreadsheets_dir.glob("*.xlsx"))

    if not xlsx_files:
        print(f"No XLSX files found in {spreadsheets_dir}")
        return

    print(f"Found {len(xlsx_files)} XLSX files")
    print("-" * 60)

    # Conversion summary
    conversion_summary = {
        "timestamp": datetime.now().isoformat(),
        "totalFiles": len(xlsx_files),
        "files": []
    }

    # Convert each file
    for xlsx_path in xlsx_files:
        filename = xlsx_path.name
        json_filename = xlsx_path.stem + ".json"
        output_path = output_dir / json_filename

        try:
            print(f"Converting: {filename}...", end=" ", flush=True)

            # Convert to JSON
            sheets_data = convert_xlsx_to_json(xlsx_path)

            # Create metadata
            json_data = {
                "source": filename,
                "converted": datetime.now().isoformat(),
                "sheets": sheets_data,
                "sheetCount": len(sheets_data)
            }

            # Write to JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False)

            # Calculate stats
            total_rows = sum(sheet["rowCount"] for sheet in sheets_data.values())

            print(f"✓ ({len(sheets_data)} sheets, {total_rows} rows)")

            conversion_summary["files"].append({
                "source": filename,
                "output": json_filename,
                "sheets": len(sheets_data),
                "rows": total_rows
            })

        except Exception as e:
            print(f"✗ ERROR: {str(e)}")
            conversion_summary["files"].append({
                "source": filename,
                "error": str(e)
            })

    # Write summary
    summary_path = output_dir / "conversion-summary.json"
    with open(summary_path, 'w', encoding='utf-8') as f:
        json.dump(conversion_summary, f, indent=2)

    print("-" * 60)
    print(f"Conversion complete!")
    print(f"Output directory: {output_dir}")
    print(f"Summary saved to: {summary_path}")

if __name__ == "__main__":
    main()
